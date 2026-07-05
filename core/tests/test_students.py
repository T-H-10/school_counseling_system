"""P0 — Student create/update validation.

Covers the two validation layers a student write passes through:
- serializer field validators (name length, ID format, phone format, the
  required-on-create enrollment fields) with their exact (Hebrew) messages, and
- the service-level duplicate-id_number guard (IntegrityError -> 400).

Messages are asserted verbatim because the React UI renders them.
"""

import io

import openpyxl
import pytest

from core.models import Student, StudentEnrollment
from core.services.student_import_export_service import (
    ExcelImportError,
    StudentImportExportService,
)
from core.tests import factories


def _valid_payload(active_year, class_level, drop=(), **overrides):
    """A valid POST /students/ body; ``drop`` removes keys, kwargs override."""
    payload = {
        "full_name": "ישראל ישראלי",
        "id_number": "123456782",
        "school_year": active_year.id,
        "class_level": class_level.id,
        "class_number": 1,
    }
    payload.update(overrides)
    for key in drop:
        payload.pop(key, None)
    return payload


# --- Happy path ------------------------------------------------------------


@pytest.mark.django_db
def test_create_student_succeeds_and_creates_enrollment(client_a, active_year, class_levels):
    payload = _valid_payload(active_year, class_levels[0])
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 201
    assert resp.data["full_name"] == "ישראל ישראלי"
    # The write-only enrollment fields produced an enrollment, reflected back.
    assert resp.data["current_class_level"] == class_levels[0].name
    assert resp.data["current_class_number"] == 1

    student = Student.objects.get(id=resp.data["id"])
    assert StudentEnrollment.objects.filter(student=student, school_year=active_year).count() == 1


@pytest.mark.django_db
@pytest.mark.parametrize("id_number", ["12345674", "123456782"])
def test_create_accepts_8_and_9_digit_ids(client_a, active_year, class_levels, id_number):
    payload = _valid_payload(active_year, class_levels[0], id_number=id_number)
    assert client_a.post("/students/", payload, format="json").status_code == 201


@pytest.mark.django_db
def test_update_student_name_persists(client_a, active_year, class_levels):
    created = client_a.post(
        "/students/", _valid_payload(active_year, class_levels[0]), format="json"
    )
    student_id = created.data["id"]

    resp = client_a.patch(f"/students/{student_id}/", {"full_name": "שם מעודכן"}, format="json")
    assert resp.status_code == 200
    assert resp.data["full_name"] == "שם מעודכן"


# --- Field validation ------------------------------------------------------


@pytest.mark.django_db
@pytest.mark.parametrize("bad_id", ["abcdefghi", "1234567", "12.345678"])
def test_create_invalid_id_number_rejected(client_a, active_year, class_levels, bad_id):
    """Non-digit or wrong-length (but <=9 char) IDs reach validate_id_number."""
    payload = _valid_payload(active_year, class_levels[0], id_number=bad_id)
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "מספר תעודת זהות לא תקין" in resp.data["id_number"]


@pytest.mark.django_db
@pytest.mark.parametrize("bad_id", ["123456789", "12345678"])
def test_create_wrong_check_digit_rejected(client_a, active_year, class_levels, bad_id):
    """Structurally valid (8-9 digits) but failing the Luhn check digit."""
    payload = _valid_payload(active_year, class_levels[0], id_number=bad_id)
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "מספר תעודת זהות לא תקין" in resp.data["id_number"]


@pytest.mark.django_db
def test_create_id_number_over_9_chars_rejected(client_a, active_year, class_levels):
    """A 10-digit ID fails the length check in validate_id_number (wrong length)."""
    payload = _valid_payload(active_year, class_levels[0], id_number="1234567890")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "מספר תעודת זהות לא תקין" in resp.data["id_number"]


@pytest.mark.django_db
def test_create_short_name_rejected(client_a, active_year, class_levels):
    payload = _valid_payload(active_year, class_levels[0], full_name="א")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "Name too short" in resp.data["full_name"]


@pytest.mark.django_db
@pytest.mark.parametrize("field", ["mother_phone", "father_phone"])
def test_create_invalid_phone_rejected(client_a, active_year, class_levels, field):
    payload = _valid_payload(active_year, class_levels[0], **{field: "123"})
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "Invalid phone number" in resp.data[field]


# --- Required-on-create enrollment fields ----------------------------------


@pytest.mark.django_db
@pytest.mark.parametrize("field", ["school_year", "class_level", "class_number"])
def test_create_missing_enrollment_field_rejected(client_a, active_year, class_levels, field):
    payload = _valid_payload(active_year, class_levels[0], drop=[field])
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "שדה זה הוא חובה" in resp.data[field]


@pytest.mark.django_db
def test_create_class_number_below_one_rejected(client_a, active_year, class_levels):
    payload = _valid_payload(active_year, class_levels[0], class_number=0)
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "class_number" in resp.data


# --- Duplicate id_number (service-level IntegrityError guard) ---------------


@pytest.mark.django_db
def test_create_duplicate_id_number_same_school_rejected(
    client_a, school_a, active_year, class_levels
):
    """A live duplicate is caught by DRF's auto UniqueValidator (code 'unique')
    at serializer time — the service's custom message is NOT what the API returns
    here (see the soft-delete edge test for where that message actually fires)."""
    factories.StudentFactory(school=school_a, id_number="123456782")
    payload = _valid_payload(active_year, class_levels[0], id_number="123456782")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert resp.data["id_number"][0].code == "unique"


@pytest.mark.django_db
def test_create_duplicate_id_number_across_schools_rejected(
    client_a, school_b, active_year, class_levels
):
    """id_number is globally unique, so a clash with ANOTHER school's student is
    still rejected (code 'unique')."""
    factories.StudentFactory(school=school_b, id_number="123456782")
    payload = _valid_payload(active_year, class_levels[0], id_number="123456782")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert resp.data["id_number"][0].code == "unique"


@pytest.mark.django_db
def test_create_duplicate_of_soft_deleted_id_hits_service_message(
    client_a, school_a, active_year, class_levels
):
    """The service's custom Hebrew duplicate message is reachable ONLY here:
    the alive-only manager hides a soft-deleted student from UniqueValidator, so
    validation passes, but the DB unique constraint (spanning soft-deleted rows)
    raises IntegrityError, which the service maps to its custom 400."""
    student = factories.StudentFactory(school=school_a, id_number="123456782")
    student.delete()  # soft-delete: sets deleted_at, row remains in the DB

    payload = _valid_payload(active_year, class_levels[0], id_number="123456782")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "תלמיד עם תעודת זהות זו כבר קיים בבית הספר" in resp.data["id_number"]


# --- Update edge: enrollment fields are ignored on update ------------------


@pytest.mark.django_db
def test_update_ignores_enrollment_fields(client_a, active_year, class_levels):
    created = client_a.post(
        "/students/", _valid_payload(active_year, class_levels[0]), format="json"
    )
    student_id = created.data["id"]

    # PATCH carrying enrollment fields must not create a second enrollment.
    resp = client_a.patch(
        f"/students/{student_id}/",
        {"class_number": 9, "class_level": class_levels[1].id},
        format="json",
    )
    assert resp.status_code == 200
    assert StudentEnrollment.objects.filter(student_id=student_id).count() == 1


# --- parents_status and notes fields ----------------------------------------


@pytest.mark.django_db
def test_create_student_with_parents_status_and_notes(client_a, active_year, class_levels):
    payload = _valid_payload(
        active_year, class_levels[0], parents_status="divorced", notes="הערה כלשהי"
    )
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 201
    assert resp.data["parents_status"] == "divorced"
    assert resp.data["notes"] == "הערה כלשהי"


@pytest.mark.django_db
def test_update_parents_status_and_notes(client_a, active_year, class_levels):
    created = client_a.post(
        "/students/", _valid_payload(active_year, class_levels[0]), format="json"
    )
    student_id = created.data["id"]

    resp = client_a.patch(
        f"/students/{student_id}/",
        {"parents_status": "married", "notes": "הערה מעודכנת"},
        format="json",
    )
    assert resp.status_code == 200
    assert resp.data["parents_status"] == "married"
    assert resp.data["notes"] == "הערה מעודכנת"


@pytest.mark.django_db
def test_invalid_parents_status_rejected(client_a, active_year, class_levels):
    payload = _valid_payload(active_year, class_levels[0], parents_status="invalid_value")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "parents_status" in resp.data


@pytest.mark.django_db
def test_parents_status_and_notes_are_optional(client_a, active_year, class_levels):
    resp = client_a.post("/students/", _valid_payload(active_year, class_levels[0]), format="json")

    assert resp.status_code == 201
    assert resp.data["parents_status"] == ""
    assert resp.data["notes"] == ""


# --- import/export -----------------------------------------------------------


def _make_xlsx(headers, rows):
    """Build an in-memory xlsx with the given headers and data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.mark.django_db
def test_export_includes_new_fields(counselor_a, school_a, active_year, class_levels):
    factories.StudentEnrollmentFactory(
        student=factories.StudentFactory(school=school_a, parents_status="divorced", notes="בדיקה"),
        school_year=active_year,
        class_level=class_levels[0],
        school=school_a,
    )
    xlsx_bytes = StudentImportExportService.export_students(school_a.students.all())

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    headers = [cell.value for cell in wb.active[1]]

    assert "מצב משפחתי" in headers
    assert "הערות" in headers

    row = list(wb.active.iter_rows(min_row=2, values_only=True))[0]
    col_idx = {h: i for i, h in enumerate(headers)}
    assert row[col_idx["מצב משפחתי"]] == "גרושים"
    assert row[col_idx["הערות"]] == "בדיקה"


@pytest.mark.django_db
def test_import_parents_status_invalid_value_rejected(counselor_a, active_year, class_levels):
    headers = [
        "שם מלא", "מספר זהות", "כיתה", "מספר כיתה", "שנת לימודים", "מצב משפחתי",
    ]
    rows = [["ישראל ישראלי", "123456782", "א", 1, active_year.name, "ערך לא חוקי"]]
    file = _make_xlsx(headers, rows)

    result = StudentImportExportService.import_students(counselor_a.user, file)

    assert result["created"] == 0
    assert len(result["errors"]) == 1
    assert "מצב משפחתי" in result["errors"][0]["message"]


@pytest.mark.django_db
def test_import_with_new_fields_succeeds(counselor_a, active_year, class_levels):
    headers = [
        "שם מלא", "מספר זהות", "כיתה", "מספר כיתה", "שנת לימודים", "מצב משפחתי", "הערות",
    ]
    rows = [["ישראל ישראלי", "123456782", "א", 1, active_year.name, "גרושים", "הערת בדיקה"]]
    file = _make_xlsx(headers, rows)

    result = StudentImportExportService.import_students(counselor_a.user, file)

    assert result["created"] == 1
    assert result["errors"] == []
    student = Student.objects.get(id_number="123456782")
    assert student.parents_status == "divorced"
    assert student.notes == "הערת בדיקה"


# --- birth_date, gender, guardian, external_care, follow_up_level -----------


@pytest.mark.django_db
def test_create_student_with_new_profile_fields(client_a, active_year, class_levels):
    payload = _valid_payload(
        active_year,
        class_levels[0],
        birth_date="2015-03-01",
        gender="male",
        guardian_name="דוד כהן",
        guardian_relation="דוד",
        guardian_phone="0501234567",
        external_care="מטופל אצל פסיכולוג בית הספר",
        follow_up_level="monitoring",
    )
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 201
    assert resp.data["birth_date"] == "2015-03-01"
    assert resp.data["gender"] == "male"
    assert resp.data["guardian_name"] == "דוד כהן"
    assert resp.data["guardian_relation"] == "דוד"
    assert resp.data["guardian_phone"] == "0501234567"
    assert resp.data["external_care"] == "מטופל אצל פסיכולוג בית הספר"
    assert resp.data["follow_up_level"] == "monitoring"


@pytest.mark.django_db
def test_update_new_profile_fields(client_a, active_year, class_levels):
    created = client_a.post(
        "/students/", _valid_payload(active_year, class_levels[0]), format="json"
    )
    student_id = created.data["id"]

    resp = client_a.patch(
        f"/students/{student_id}/",
        {"follow_up_level": "at_risk", "gender": "female"},
        format="json",
    )
    assert resp.status_code == 200
    assert resp.data["follow_up_level"] == "at_risk"
    assert resp.data["gender"] == "female"


@pytest.mark.django_db
def test_invalid_gender_rejected(client_a, active_year, class_levels):
    payload = _valid_payload(active_year, class_levels[0], gender="invalid_value")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "gender" in resp.data


@pytest.mark.django_db
def test_invalid_follow_up_level_rejected(client_a, active_year, class_levels):
    payload = _valid_payload(active_year, class_levels[0], follow_up_level="invalid_value")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "follow_up_level" in resp.data


@pytest.mark.django_db
def test_invalid_guardian_phone_rejected(client_a, active_year, class_levels):
    payload = _valid_payload(active_year, class_levels[0], guardian_phone="123")
    resp = client_a.post("/students/", payload, format="json")

    assert resp.status_code == 400
    assert "guardian_phone" in resp.data


@pytest.mark.django_db
def test_new_profile_fields_are_optional(client_a, active_year, class_levels):
    resp = client_a.post("/students/", _valid_payload(active_year, class_levels[0]), format="json")

    assert resp.status_code == 201
    assert resp.data["birth_date"] is None
    assert resp.data["gender"] == ""
    assert resp.data["guardian_name"] == ""
    assert resp.data["guardian_relation"] == ""
    assert resp.data["guardian_phone"] == ""
    assert resp.data["external_care"] == ""
    assert resp.data["follow_up_level"] == "none"


@pytest.mark.django_db
def test_export_includes_new_profile_fields(counselor_a, school_a, active_year, class_levels):
    factories.StudentEnrollmentFactory(
        student=factories.StudentFactory(
            school=school_a,
            birth_date="2015-03-01",
            gender="male",
            guardian_name="דוד כהן",
            guardian_relation="דוד",
            guardian_phone="0501234567",
            external_care="מטופל אצל פסיכולוג",
            follow_up_level="at_risk",
        ),
        school_year=active_year,
        class_level=class_levels[0],
        school=school_a,
    )
    xlsx_bytes = StudentImportExportService.export_students(school_a.students.all())

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    headers = [cell.value for cell in wb.active[1]]
    row = list(wb.active.iter_rows(min_row=2, values_only=True))[0]
    col_idx = {h: i for i, h in enumerate(headers)}

    assert row[col_idx["תאריך לידה"]] == "2015-03-01"
    assert row[col_idx["מגדר"]] == "זכר"
    assert row[col_idx["שם אפוטרופוס"]] == "דוד כהן"
    assert row[col_idx["קרבת אפוטרופוס"]] == "דוד"
    assert row[col_idx["טלפון אפוטרופוס"]] == "0501234567"
    assert row[col_idx["גורמים מטפלים"]] == "מטופל אצל פסיכולוג"
    assert row[col_idx["רמת מעקב"]] == "בסיכון"


@pytest.mark.django_db
def test_import_gender_invalid_value_rejected(counselor_a, active_year, class_levels):
    headers = ["שם מלא", "מספר זהות", "כיתה", "מספר כיתה", "שנת לימודים", "מגדר"]
    rows = [["ישראל ישראלי", "123456782", "א", 1, active_year.name, "ערך לא חוקי"]]
    file = _make_xlsx(headers, rows)

    result = StudentImportExportService.import_students(counselor_a.user, file)

    assert result["created"] == 0
    assert len(result["errors"]) == 1
    assert "מגדר" in result["errors"][0]["message"]


@pytest.mark.django_db
def test_import_follow_up_level_invalid_value_rejected(counselor_a, active_year, class_levels):
    headers = ["שם מלא", "מספר זהות", "כיתה", "מספר כיתה", "שנת לימודים", "רמת מעקב"]
    rows = [["ישראל ישראלי", "123456782", "א", 1, active_year.name, "ערך לא חוקי"]]
    file = _make_xlsx(headers, rows)

    result = StudentImportExportService.import_students(counselor_a.user, file)

    assert result["created"] == 0
    assert len(result["errors"]) == 1
    assert "רמת מעקב" in result["errors"][0]["message"]


@pytest.mark.django_db
def test_import_birth_date_invalid_value_rejected(counselor_a, active_year, class_levels):
    headers = ["שם מלא", "מספר זהות", "כיתה", "מספר כיתה", "שנת לימודים", "תאריך לידה"]
    rows = [["ישראל ישראלי", "123456782", "א", 1, active_year.name, "לא-תאריך"]]
    file = _make_xlsx(headers, rows)

    result = StudentImportExportService.import_students(counselor_a.user, file)

    assert result["created"] == 0
    assert len(result["errors"]) == 1
    assert "תאריך לידה" in result["errors"][0]["message"]


@pytest.mark.django_db
def test_import_with_all_new_fields_succeeds(counselor_a, active_year, class_levels):
    headers = [
        "שם מלא", "מספר זהות", "כיתה", "מספר כיתה", "שנת לימודים",
        "תאריך לידה", "מגדר", "שם אפוטרופוס", "קרבת אפוטרופוס",
        "טלפון אפוטרופוס", "גורמים מטפלים", "רמת מעקב",
    ]
    rows = [[
        "ישראל ישראלי", "123456782", "א", 1, active_year.name,
        "2015-03-01", "זכר", "דוד כהן", "דוד",
        "0501234567", "מטופל אצל פסיכולוג", "בסיכון",
    ]]
    file = _make_xlsx(headers, rows)

    result = StudentImportExportService.import_students(counselor_a.user, file)

    assert result["created"] == 1
    assert result["errors"] == []
    student = Student.objects.get(id_number="123456782")
    assert student.birth_date.isoformat() == "2015-03-01"
    assert student.gender == "male"
    assert student.guardian_name == "דוד כהן"
    assert student.guardian_relation == "דוד"
    assert student.guardian_phone == "0501234567"
    assert student.external_care == "מטופל אצל פסיכולוג"
    assert student.follow_up_level == "at_risk"
