"""Excel import/export for students.

Holds the workbook parsing/validation and workbook-building logic that used to
live inline in ``StudentViewSet``. Behavior is unchanged: the same column map,
the same Hebrew validation messages, and the same result shape are produced.
HTTP concerns (status codes, file responses) stay in the view.
"""

import datetime
import io

import openpyxl
from core.models import ClassLevel, SchoolYear
from core.services.student_service import StudentService


class ExcelImportError(Exception):
    """Raised for whole-file problems the view should surface as HTTP 400."""


# Mirrors Student.PARENTS_STATUS_CHOICES — update both together if choices change.
_PARENTS_STATUS_HEBREW_TO_VALUE = {
    "נשואים": "married",
    "גרושים": "divorced",
    "פרודים": "separated",
    "חד הוריות": "single_parent",
    "שכול": "widowed",
    "אחר": "other",
}
_PARENTS_STATUS_VALUE_TO_HEBREW = {v: k for k, v in _PARENTS_STATUS_HEBREW_TO_VALUE.items()}

# Mirrors Student.GENDER_CHOICES — update both together if choices change.
_GENDER_HEBREW_TO_VALUE = {
    "זכר": "male",
    "נקבה": "female",
}
_GENDER_VALUE_TO_HEBREW = {v: k for k, v in _GENDER_HEBREW_TO_VALUE.items()}

# Mirrors Student.FOLLOW_UP_CHOICES — update both together if choices change.
_FOLLOW_UP_HEBREW_TO_VALUE = {
    "רגיל": "none",
    "במעקב": "monitoring",
    "בסיכון": "at_risk",
}
_FOLLOW_UP_VALUE_TO_HEBREW = {v: k for k, v in _FOLLOW_UP_HEBREW_TO_VALUE.items()}

COL_MAP = {
    "שם מלא": "full_name",
    "מספר זהות": "id_number",
    "כיתה": "class_level_name",
    "מספר כיתה": "class_number",
    "שנת לימודים": "school_year_name",
    "שם אם": "mother_name",
    "טלפון אם": "mother_phone",
    "שם אב": "father_name",
    "טלפון אב": "father_phone",
    "כתובת": "address",
    "מצב משפחתי": "parents_status",
    "הערות": "notes",
    "תאריך לידה": "birth_date",
    "מגדר": "gender",
    "שם אפוטרופוס": "guardian_name",
    "קרבת אפוטרופוס": "guardian_relation",
    "טלפון אפוטרופוס": "guardian_phone",
    "גורמים מטפלים": "external_care",
    "רמת מעקב": "follow_up_level",
}
REQUIRED_COLS = {
    "full_name",
    "id_number",
    "class_level_name",
    "class_number",
    "school_year_name",
}
REQUIRED_HEB = {
    "full_name": "שם מלא",
    "id_number": "מספר זהות",
    "class_level_name": "כיתה",
    "class_number": "מספר כיתה",
    "school_year_name": "שנת לימודים",
}

EXPORT_HEADERS = [
    "שם מלא",
    "מספר זהות",
    "כיתה",
    "מספר כיתה",
    "שנת לימודים",
    "שם אם",
    "טלפון אם",
    "שם אב",
    "טלפון אב",
    "כתובת",
    "מצב משפחתי",
    "הערות",
    "תאריך לידה",
    "מגדר",
    "שם אפוטרופוס",
    "קרבת אפוטרופוס",
    "טלפון אפוטרופוס",
    "גורמים מטפלים",
    "רמת מעקב",
]


class StudentImportExportService:
    @staticmethod
    def import_students(user, file):
        """Parse an uploaded .xlsx and bulk-create students.

        Returns a ``{'created': int, 'errors': [...]}`` dict on success.
        Raises ``ExcelImportError`` for an unreadable file or missing
        required columns (the view turns these into HTTP 400).
        """
        # Magic-byte pre-check: a real .xlsx is a ZIP container. Rejecting
        # renamed binaries here means openpyxl only ever sees ZIP files; the
        # except below still catches ZIPs that aren't valid Excel inside.
        head = file.read(4)
        file.seek(0)
        if head != b"PK\x03\x04":
            raise ExcelImportError("קובץ Excel לא תקין. ודא שהקובץ הוא מסוג .xlsx")

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            raise ExcelImportError("קובץ Excel לא תקין. ודא שהקובץ הוא מסוג .xlsx")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"created": 0, "errors": []}

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        col_indices = {COL_MAP[h]: idx for idx, h in enumerate(headers) if h in COL_MAP}

        # Fail fast if any required column is absent from the header row
        missing_cols = [REQUIRED_HEB[f] for f in REQUIRED_COLS if f not in col_indices]
        if missing_cols:
            raise ExcelImportError(
                f"עמודות חסרות בקובץ: {', '.join(missing_cols)}. "
                "עמודות חובה בשורת הכותרת: שם מלא, מספר זהות, כיתה, מספר כיתה, שנת לימודים"
            )

        class_levels = {cl.name: cl for cl in ClassLevel.objects.all()}
        school_years = {sy.name: sy for sy in SchoolYear.objects.all()}

        def cell(row, field):
            idx = col_indices.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            val = row[idx]
            return (
                str(int(val)) if isinstance(val, float) and val.is_integer() else str(val).strip()
            )

        def cell_date(row, field):
            idx = col_indices.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            val = row[idx]
            if isinstance(val, datetime.datetime):
                return val.date().isoformat()
            if isinstance(val, datetime.date):
                return val.isoformat()
            return str(val).strip()

        parsed = []
        pre_errors = []

        for row_num, row in enumerate(rows[1:], start=2):
            # Skip fully empty rows (trailing blank lines in Excel)
            if all(c is None for c in row):
                continue

            cl_name = cell(row, "class_level_name")
            sy_name = cell(row, "school_year_name")
            cn_raw = cell(row, "class_number")
            try:
                class_number = int(float(cn_raw)) if cn_raw else 0
            except (ValueError, TypeError):
                class_number = 0

            full_name = cell(row, "full_name") or ""
            id_number = cell(row, "id_number") or ""
            class_level = class_levels.get(cl_name) if cl_name else None
            school_year = school_years.get(sy_name) if sy_name else None

            # Per-row structural validation with specific Hebrew messages
            row_errors = []
            if not full_name:
                row_errors.append("שם מלא חסר")
            if not id_number:
                row_errors.append("מספר זהות חסר")
            if not school_year:
                label = f'"{sy_name}"' if sy_name else "ריק"
                row_errors.append(f"שנת לימודים {label} לא קיימת במערכת")
            if not class_level:
                label = f'"{cl_name}"' if cl_name else "ריק"
                row_errors.append(f"כיתה {label} לא קיימת במערכת (ערכים אפשריים: א–ח)")
            if not class_number:
                row_errors.append("מספר כיתה חסר או לא תקין")

            parents_status_raw = cell(row, "parents_status")
            parents_status = ""
            if parents_status_raw:
                resolved = _PARENTS_STATUS_HEBREW_TO_VALUE.get(parents_status_raw)
                if resolved is None:
                    valid = "، ".join(_PARENTS_STATUS_HEBREW_TO_VALUE)
                    row_errors.append(
                        f'מצב משפחתי "{parents_status_raw}" אינו חוקי. ערכים אפשריים: {valid}'
                    )
                else:
                    parents_status = resolved

            gender_raw = cell(row, "gender")
            gender = ""
            if gender_raw:
                resolved = _GENDER_HEBREW_TO_VALUE.get(gender_raw)
                if resolved is None:
                    valid = "، ".join(_GENDER_HEBREW_TO_VALUE)
                    row_errors.append(f'מגדר "{gender_raw}" אינו חוקי. ערכים אפשריים: {valid}')
                else:
                    gender = resolved

            follow_up_raw = cell(row, "follow_up_level")
            follow_up_level = "none"
            if follow_up_raw:
                resolved = _FOLLOW_UP_HEBREW_TO_VALUE.get(follow_up_raw)
                if resolved is None:
                    valid = "، ".join(_FOLLOW_UP_HEBREW_TO_VALUE)
                    row_errors.append(
                        f'רמת מעקב "{follow_up_raw}" אינה חוקית. ערכים אפשריים: {valid}'
                    )
                else:
                    follow_up_level = resolved

            birth_date_raw = cell_date(row, "birth_date")
            birth_date = None
            if birth_date_raw:
                try:
                    birth_date = datetime.date.fromisoformat(birth_date_raw)
                except ValueError:
                    row_errors.append(
                        f'תאריך לידה "{birth_date_raw}" אינו תקין. פורמט נדרש: YYYY-MM-DD'
                    )

            if row_errors:
                pre_errors.append({"row": row_num, "message": " | ".join(row_errors)})
                continue

            data = {
                "full_name": full_name,
                "id_number": id_number,
                "address": cell(row, "address"),
                "mother_name": cell(row, "mother_name"),
                "mother_phone": cell(row, "mother_phone"),
                "father_name": cell(row, "father_name"),
                "father_phone": cell(row, "father_phone"),
                "parents_status": parents_status,
                "notes": cell(row, "notes") or "",
                "birth_date": birth_date,
                "gender": gender,
                "guardian_name": cell(row, "guardian_name") or "",
                "guardian_relation": cell(row, "guardian_relation") or "",
                "guardian_phone": cell(row, "guardian_phone") or "",
                "external_care": cell(row, "external_care") or "",
                "follow_up_level": follow_up_level,
                "class_number": class_number,
                "class_level": class_level,
                "school_year": school_year,
            }
            parsed.append((row_num, data))

        result = StudentService.bulk_create_students(user, parsed)
        result["errors"] = pre_errors + result["errors"]
        return result

    @staticmethod
    def export_students(queryset):
        """Build an .xlsx of the given students and return its bytes."""
        students = queryset.prefetch_related(
            "enrollments__class_level", "enrollments__school_year"
        ).order_by("full_name")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "תלמידים"
        ws.sheet_view.rightToLeft = True
        ws.append(EXPORT_HEADERS)

        for student in students:
            enrollment = student.enrollments.order_by("-created_at").first()
            ws.append(
                [
                    student.full_name,
                    student.id_number,
                    enrollment.class_level.name if enrollment and enrollment.class_level else "",
                    enrollment.class_number if enrollment else "",
                    enrollment.school_year.name if enrollment and enrollment.school_year else "",
                    student.mother_name or "",
                    student.mother_phone or "",
                    student.father_name or "",
                    student.father_phone or "",
                    student.address or "",
                    _PARENTS_STATUS_VALUE_TO_HEBREW.get(student.parents_status, ""),
                    student.notes or "",
                    student.birth_date.isoformat() if student.birth_date else "",
                    _GENDER_VALUE_TO_HEBREW.get(student.gender, ""),
                    student.guardian_name or "",
                    student.guardian_relation or "",
                    student.guardian_phone or "",
                    student.external_care or "",
                    _FOLLOW_UP_VALUE_TO_HEBREW.get(student.follow_up_level, ""),
                ]
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
