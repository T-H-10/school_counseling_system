from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.parsers import MultiPartParser
from django_filters.rest_framework import DjangoFilterBackend
import io
import openpyxl

from core.services.dashboard_service import DashboardService
from core.filters import StudentFilter

from .services.student_timeline_service import StudentTimelineService
from .services.student_service import StudentService
from .services.student_enrollment_service import StudentEnrollmentService
from .services.student_event_service import StudentEventService
from .services.class_session_service import ClassSessionService
from .services.school_service import SchoolService
from .services.class_level_service import ClassLevelService
from .services.school_year_service import SchoolYearService
from .services.counselor_service import CounselorService

from .models import Student, StudentEnrollment, StudentEvent, ClassSession, School, ClassLevel, SchoolYear, Counselor
from .serializers import StudentSerializer, StudentEnrollmentSerializer, StudentEventSerializer, ClassSessionSerializer, SchoolSerializer, ClassLevelSerializer, SchoolYearSerializer, CounselorSerializer
from .permissions import IsCounselor

class BaseSchoolViewSet(ModelViewSet):
    model = None
    
    def get_queryset(self):

        if getattr(self, "swagger_fake_view", False):
            return self.model.objects.none()
        
        user = self.request.user

        if not user.is_authenticated or not hasattr(user, "counselor"):
            return self.model.objects.none()
        
        school = user.counselor.school

        return self.model.objects.filter(school=school)

class StudentViewSet(BaseSchoolViewSet):
    permission_classes = [IsCounselor]
    model = Student
    serializer_class = StudentSerializer

    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = StudentFilter

    search_fields = ["full_name", "id_number"]
    ordering_fields = ["full_name", "id_number", "created_at"]
    ordering = ["full_name"]

    def get_queryset(self):
        return super().get_queryset().distinct()

    @action(detail=True, methods=["get"])
    def timeline(self, request, pk=None):
        student = self.get_object()
        data = StudentTimelineService.get_timeline(student)
        return Response(
            {
                "student_id": student.id,
                "timeline": data
            }
        )

    @action(detail=False, methods=['post'], url_path='import', parser_classes=[MultiPartParser])
    def import_students(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'לא נשלח קובץ'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'error': 'קובץ Excel לא תקין. ודא שהקובץ הוא מסוג .xlsx'}, status=400)

        COL_MAP = {
            'שם מלא': 'full_name',
            'מספר זהות': 'id_number',
            'כיתה': 'class_level_name',
            'מספר כיתה': 'class_number',
            'שנת לימודים': 'school_year_name',
            'שם אם': 'mother_name',
            'טלפון אם': 'mother_phone',
            'שם אב': 'father_name',
            'טלפון אב': 'father_phone',
            'כתובת': 'address',
        }
        REQUIRED_COLS = {'full_name', 'id_number', 'class_level_name', 'class_number', 'school_year_name'}
        REQUIRED_HEB  = {'full_name': 'שם מלא', 'id_number': 'מספר זהות',
                         'class_level_name': 'כיתה', 'class_number': 'מספר כיתה',
                         'school_year_name': 'שנת לימודים'}

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({'created': 0, 'errors': []})

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        col_indices = {COL_MAP[h]: idx for idx, h in enumerate(headers) if h in COL_MAP}

        # Fail fast if any required column is absent from the header row
        missing_cols = [REQUIRED_HEB[f] for f in REQUIRED_COLS if f not in col_indices]
        if missing_cols:
            return Response({
                'error': (
                    f'עמודות חסרות בקובץ: {", ".join(missing_cols)}. '
                    'עמודות חובה בשורת הכותרת: שם מלא, מספר זהות, כיתה, מספר כיתה, שנת לימודים'
                )
            }, status=400)

        class_levels = {cl.name: cl for cl in ClassLevel.objects.all()}
        school_years = {sy.name: sy for sy in SchoolYear.objects.all()}

        def cell(row, field):
            idx = col_indices.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            val = row[idx]
            return str(int(val)) if isinstance(val, float) and val.is_integer() else str(val).strip()

        parsed = []
        pre_errors = []

        for row_num, row in enumerate(rows[1:], start=2):
            # Skip fully empty rows (trailing blank lines in Excel)
            if all(c is None for c in row):
                continue

            cl_name = cell(row, 'class_level_name')
            sy_name = cell(row, 'school_year_name')
            cn_raw  = cell(row, 'class_number')
            try:
                class_number = int(float(cn_raw)) if cn_raw else 0
            except (ValueError, TypeError):
                class_number = 0

            full_name   = cell(row, 'full_name') or ''
            id_number   = cell(row, 'id_number') or ''
            class_level = class_levels.get(cl_name) if cl_name else None
            school_year = school_years.get(sy_name) if sy_name else None

            # Per-row structural validation with specific Hebrew messages
            row_errors = []
            if not full_name:
                row_errors.append('שם מלא חסר')
            if not id_number:
                row_errors.append('מספר זהות חסר')
            if not school_year:
                label = f'"{sy_name}"' if sy_name else 'ריק'
                row_errors.append(f'שנת לימודים {label} לא קיימת במערכת')
            if not class_level:
                label = f'"{cl_name}"' if cl_name else 'ריק'
                row_errors.append(f'כיתה {label} לא קיימת במערכת (ערכים אפשריים: א–ח)')
            if not class_number:
                row_errors.append('מספר כיתה חסר או לא תקין')

            if row_errors:
                pre_errors.append({'row': row_num, 'message': ' | '.join(row_errors)})
                continue

            data = {
                'full_name':    full_name,
                'id_number':    id_number,
                'address':      cell(row, 'address'),
                'mother_name':  cell(row, 'mother_name'),
                'mother_phone': cell(row, 'mother_phone'),
                'father_name':  cell(row, 'father_name'),
                'father_phone': cell(row, 'father_phone'),
                'class_number': class_number,
                'class_level':  class_level,
                'school_year':  school_year,
            }
            parsed.append((row_num, data))

        result = StudentService.bulk_create_students(request.user, parsed)
        result['errors'] = pre_errors + result['errors']
        return Response(result)

    @action(detail=False, methods=['get'], url_path='export')
    def export_students(self, request):
        students = (
            self.get_queryset()
            .prefetch_related('enrollments__class_level', 'enrollments__school_year')
            .order_by('full_name')
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'תלמידים'
        ws.sheet_view.rightToLeft = True
        ws.append(['שם מלא', 'מספר זהות', 'כיתה', 'מספר כיתה', 'שנת לימודים',
                   'שם אם', 'טלפון אם', 'שם אב', 'טלפון אב', 'כתובת'])

        for student in students:
            enrollment = student.enrollments.order_by('-created_at').first()
            ws.append([
                student.full_name,
                student.id_number,
                enrollment.class_level.name if enrollment and enrollment.class_level else '',
                enrollment.class_number if enrollment else '',
                enrollment.school_year.name if enrollment and enrollment.school_year else '',
                student.mother_name or '',
                student.mother_phone or '',
                student.father_name or '',
                student.father_phone or '',
                student.address or '',
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
        return response

    def perform_create(self, serializer):
        student = StudentService.create_student(
            self.request.user,
            serializer.validated_data
        )
        serializer.instance = student

    def perform_update(self, serializer):
        student = StudentService.update_student(
            self.request.user,
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = student

    def perform_destroy(self, instance):
        StudentService.delete_student(
            self.request.user,
            instance
        )


class StudentEnrollmentViewSet(BaseSchoolViewSet):
    permission_classes = [IsCounselor]
    model = StudentEnrollment
    serializer_class = StudentEnrollmentSerializer

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['student']

    def perform_create(self, serializer):
        enrollment = StudentEnrollmentService.create_enrollment(
            self.request.user,
            serializer.validated_data
        )
        serializer.instance = enrollment

    def perform_update(self, serializer):
        enrollment = StudentEnrollmentService.update_enrollment(
            self.request.user,
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = enrollment

    def perform_destroy(self, instance):
        StudentEnrollmentService.delete_enrollment(
            self.request.user,
            instance
        ) 


class StudentEventViewSet(BaseSchoolViewSet):
    permission_classes = [IsCounselor]
    model = StudentEvent
    serializer_class = StudentEventSerializer
   
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['student']

    def perform_create(self, serializer):
        event = StudentEventService.create_event(
            self.request.user,
            serializer.validated_data
        )
        serializer.instance = event

    def perform_update(self, serializer):
        event = StudentEventService.update_event(
            self.request.user,
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = event

    def perform_destroy(self, instance):
        StudentEventService.delete_event(
            self.request.user,
            instance
        )        


class ClassSessionViewSet(BaseSchoolViewSet):
    permission_classes = [IsCounselor]
    model = ClassSession
    serializer_class = ClassSessionSerializer

    def perform_create(self, serializer):
        session = ClassSessionService.create_session(
            self.request.user,
            serializer.validated_data
        )
        serializer.instance = session

    def perform_update(self, serializer):
        session = ClassSessionService.update_session(
            self.request.user,
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = session

    def perform_destroy(self, instance):
        ClassSessionService.delete_session(
            self.request.user,
            instance
        )

    @action(detail=False, methods=["get"])
    def calendar(self, request):
        start = request.query_params.get("start")
        end = request.query_params.get("end")

        data = ClassSessionService.get_calendar(
            request.user, start, end
        )
        return Response(data)

class SchoolViewSet(ModelViewSet):
    permission_classes = [IsAdminUser]
    queryset = School.objects.all()
    serializer_class = SchoolSerializer

    def perform_create(self, serializer):
        school = SchoolService.create_school(serializer.validated_data)
        serializer.instance = school

    def perform_update(self, serializer):
        school = SchoolService.update_school(
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = school

    def perform_destroy(self, instance):
        SchoolService.delete_school(instance)        


class ClassLevelViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = ClassLevel.objects.all()
    serializer_class = ClassLevelSerializer

    def perform_create(self, serializer):
        level = ClassLevelService.create_class_level(serializer.validated_data)
        serializer.instance = level

    def perform_update(self, serializer):
        level = ClassLevelService.update_class_level(
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = level

    def perform_destroy(self, instance):
        ClassLevelService.delete_class_level(instance)


class SchoolYearViewSet(ModelViewSet):
    queryset = SchoolYear.objects.all()
    serializer_class = SchoolYearSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAdminUser()]

    def perform_create(self, serializer):
        year = SchoolYearService.create_school_year(serializer.validated_data)
        serializer.instance = year

    def perform_update(self, serializer):
        year = SchoolYearService.update_school_year(
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = year

    def perform_destroy(self, instance):
        SchoolYearService.delete_school_year(
            instance
            )


class CounselorViewSet(ModelViewSet):
    permission_classes = [IsAdminUser]
    queryset = Counselor.objects.all()
    serializer_class = CounselorSerializer

    def perform_create(self, serializer):
        counselor = CounselorService.create_counselor(
            serializer.validated_data
            )
        serializer.instance = counselor

    def perform_update(self, serializer):
        counselor = CounselorService.update_counselor(
            self.get_object(),
            serializer.validated_data
        )
        serializer.instance = counselor

    def perform_destroy(self, instance):
        CounselorService.delete_counselor(
            instance
        )
    
    @action(detail=True, methods=["post"], permission_classes=[IsAdminUser])
    def reset_password(self, request, pk = None):
        counselor = self.get_object()

        CounselorService.reset_password(
            counselor,
            request.data["new_password"]
        )

        return Response({"status": "password updated"})
    

class DashboardView(APIView):
    permission_classes = [IsCounselor]

    def get(self, request):
        data = DashboardService.get_dashboard(request.user)
        return Response(data)